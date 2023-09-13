import os
import requests
import chardet
import re
import json
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import warnings
import jieba.analyse
from wordcloud import WordCloud
from matplotlib.image import imread
from wordcloud import ImageColorGenerator
import random
import time
import openpyxl

headers = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36"
}
count = 0
words = []
content = input("请输入你要检索的内容: ")
keyword = content.encode('utf-8')
keyword = str(keyword)[2:-1].replace('\\x', '%').upper()

cookies = requests.get('https://www.bilibili.com/', headers=headers).cookies.get_dict()
bvids = []
for page in range(0, 15):
    url = f"https://api.bilibili.com/x/web-interface/wbi/search/type?keyword={keyword}&search_type=video&page={page + 1}"
    response = requests.get(url, headers=headers, cookies=cookies)
    
    bvids.extend(re.findall(r'"bvid": "(\w+)', response.text))
    print(f"正在获取页面bvid，现在已进行到 {page+1}/15")
    print(f"请求状态码：{response.status_code}")
    response_dict = json.loads(response.text)
    result_list = response_dict.get('data', {}).get('result', [])
    for result in result_list:
        bvid = result.get('bvid')
        if bvid:
            bvids.append(bvid)
            # cid = get_cid(bvid)
            # print(f"视频 {bvid} 的 cid 为：{cid}")
    
    time.sleep(round(random.uniform(0, 3), 3))
print(bvids)

# 获取视频cid
def get_cid(bvid):
    API = 'https://api.bilibili.com/x/player/pagelist'
    params = {
        'bvid': bvid,
        'jsonp': 'jsonp',
    }
    response = requests.get(url=API, headers=headers, params=params)
    html = response.text
    dict = json.loads(html)
    result = dict["data"][0]["cid"]
    return result

def get_barrage(cid):
    API = 'https://api.bilibili.com/x/v1/dm/list.so'
    params = {
        'oid': cid,
    }
    response = requests.get(url=API, headers=headers, params=params)

    # 获取网页编码类型并解码网页内容
    response.encoding = chardet.detect(response.content)['encoding']
    html = response.text
    
    pattern = re.compile(r'<d p="(.*?)">(.*?)</d>')
    result = pattern.findall(html)
    
    words = []
    count = 0

    # 将弹幕内容解析为字典，并保存到列表中
    for item in result:
        count += 1
        info = item[0].split(',')
        appear_time = info[0]
        type = info[1]
        font_size = info[2]
        font_color = info[3]
        send_time = info[4]
        pool = info[5]
        author_id = info[6]
        database_id = info[7]
        content = item[1]

        if isinstance(content, (float, int)):
            continue

        words.append({
            'atime': appear_time,
            'type': type,
            'font_size': font_size,
            'font_color': font_color,
            'send_time': send_time,
            'pool': pool,
            'author_id': author_id,
            'database_id': database_id,
            'content': content,
        })

        print('第{}条数据解析完成'.format(count))

    return words
def write_to_excel(words, filename, sheet_name='Sheet1'):
    try:
        # 创建Excel工作簿
        work_book = openpyxl.Workbook()
        # 创建Excel工作表
        sheet = work_book.active
        sheet.title = sheet_name

        head = []
        # 获取表头
        for k in words[0].keys():
            head.append(k)

        # 将表头写入Excel表格中
        for i in range(len(head)):
            sheet.cell(row=1, column=i + 1, value=head[i])

        i = 2
        # 遍历数据并写入Excel表格中
        for item in words:
            for j in range(len(head)):
                sheet.cell(row=i, column=j + 1, value=item[head[j]])
            i += 1

        # 保存Excel文件
        work_book.save(filename)
        print('写入excel成功！')

    except Exception as e:
        print('写入excel失败！', e)

def content_analysis(data):
    # 获取所有弹幕内容，将它们拼接成一个字符串
    content_list = data['content'].tolist()
    content = ''.join(content_list)

    # 统计每个弹幕的出现次数
    word_counts = {}
    for item in content_list:
        if item in word_counts:
            word_counts[item] += 1
        else:
            word_counts[item] = 1

    # 对弹幕出现次数进行排序
    sorted_word_counts = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)

    # 输出弹幕数量排名前20的弹幕
    print("弹幕数量排名前20的弹幕：")
    for i in range(min(20, len(sorted_word_counts))):
        print("{}: {}".format(sorted_word_counts[i][0], sorted_word_counts[i][1]))
    
    
    # 读入背景图像并生成颜色方案
    back_img = imread("D:/软工/海浪.jpg")
    img_colors = ImageColorGenerator(back_img)
    
    # 使用jieba库提取关键词，并生成词云图
    content_list = data['content'].tolist()
    content_text = ''.join(content_list)
    tags = jieba.analyse.extract_tags(content_text, topK=600, withWeight=True, allowPOS=('n', 'vn', 'v'))

    cloud_data = {item[0]: item[1] for item in tags}

    word_cloud = WordCloud(
        font_path="D:/软工/ChillKai_Big5.ttf",
        background_color="white",
        max_words=500,
        max_font_size=35,
        width=2500,
        mask=back_img,
        height=1080,
    ).generate_from_frequencies(cloud_data)

    sub_color = word_cloud.recolor(color_func=img_colors)

    # 绘制词云图并保存为图片文件
    plt.figure(figsize=(25, 25))
    plt.imshow(sub_color, interpolation='bilinear')
    plt.axis("off")
    plt.savefig('弹幕词云.png')

    plt.show()
    
    top20_words = dict(sorted_word_counts[:20])
    top20_list = list(top20_words.items())

    top20_df = pd.DataFrame(top20_list, columns=['弹幕内容', '出现次数'])
    top20_df.to_excel('top20弹幕.xlsx', index=False)

    top10000_words = dict(sorted_word_counts[:10000])
    top10000_list = list(top10000_words.items())

    top10000_df = pd.DataFrame(top10000_list, columns=['弹幕内容', '出现次数'])
    top10000_df.to_excel('top10000弹幕.xlsx', index=False)

def crawl_video_barrage():
    
    words = []
    count = 0

    # 遍历视频链接，并爬取每个视频的弹幕
    for bvid in bvids:
        count += 1
        cid = get_cid(bvid)
        print(cid)
        video_barrage = get_barrage(cid)
        print(video_barrage)
        words.extend(video_barrage)
        print('第{}个视频的弹幕爬取完成'.format(count))

    # 将弹幕数据写入Excel文件，并进行弹幕内容分析
    write_to_excel(words, 'danmu.xls', sheet_name='danmu')
    data = pd.read_excel('danmu.xls', sheet_name='danmu')
    content_analysis(data)

crawl_video_barrage()